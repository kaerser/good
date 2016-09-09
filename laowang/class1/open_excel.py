#!/usr/bin/python
# -*-coding:utf-8-*-
"""
@author = 'kaerser'
@time:8/28/2016 10:20 PM
"""

# from openpyxl import Workbook
import openpyxl

if __name__ == '__main__':
	# wb = Workbook()
	# ws = wb.active
	# ws1 = wb.create_sheet()
	# ws1.title = 'New Title'
	# # ws2 = wb.get_sheet_by_name('New Title')
	# wb.save('myexcel.xlsx')

	wb = openpyxl.load_workbook('myexcel.xlsx')
	print wb.get_sheet_names()